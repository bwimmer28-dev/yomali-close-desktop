"""
Output Formatting (v2)

Generates Excel reports in the legacy bridge table format:
- Summary sheet with daily status overview
- Bridge table for month-end close
- Exception breakdown by reason code
- Processor details with Proof A/B results
"""
from __future__ import annotations

import io
from datetime import date
from typing import Dict, List, Any
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    DailyStatus, ReconException, MonthEndBridge,
    ReconciliationStatus, ReasonCode
)


# =============================================================================
# Style Constants
# =============================================================================

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

CURRENCY_FORMAT = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
PERCENT_FORMAT = '0.00%'


def get_status_fill(status: ReconciliationStatus) -> PatternFill:
    """Get fill color for status"""
    if status == ReconciliationStatus.GREEN:
        return GREEN_FILL
    elif status == ReconciliationStatus.YELLOW:
        return YELLOW_FILL
    else:
        return RED_FILL


# =============================================================================
# Main Output Function
# =============================================================================

def write_recon_xlsx(
    output: io.BytesIO | Path,
    summary_df: pd.DataFrame,
    exceptions_df: pd.DataFrame,
    meta: Dict[str, Any],
) -> None:
    """
    Write reconciliation results to Excel in bridge table format.
    
    Sheets:
    - Summary: Daily status overview with traffic lights
    - Exceptions: Variance breakdown by reason code
    - Bridge: Month-end style summary table
    - Details: Raw data for drilldown
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Get daily statuses from meta
    daily_statuses = meta.get("daily_statuses", [])
    
    # Create sheets
    _create_summary_sheet(wb, daily_statuses, meta)
    _create_exceptions_sheet(wb, exceptions_df, meta)
    _create_bridge_sheet(wb, daily_statuses, meta)
    _create_details_sheet(wb, summary_df, meta)
    
    # Save
    if isinstance(output, io.BytesIO):
        wb.save(output)
        output.seek(0)
    else:
        wb.save(str(output))


# =============================================================================
# Summary Sheet
# =============================================================================

def _create_summary_sheet(wb: Workbook, daily_statuses: List[Dict], meta: Dict):
    """Create summary sheet with daily status overview"""
    ws = wb.create_sheet("Summary")
    
    # Title
    ws["A1"] = f"Daily Reconciliation Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Entity: {meta.get('entity', 'Unknown')}"
    ws["A3"] = f"Date: {meta.get('target_day', '')}"
    
    # Status summary
    summary = meta.get("summary", {})
    ws["A5"] = "Status Overview"
    ws["A5"].font = Font(bold=True)
    
    row = 6
    ws[f"A{row}"] = "Total Processors:"
    ws[f"B{row}"] = summary.get("total_processors", 0)
    row += 1
    ws[f"A{row}"] = "Green (OK):"
    ws[f"B{row}"] = summary.get("green_count", 0)
    ws[f"B{row}"].fill = GREEN_FILL
    row += 1
    ws[f"A{row}"] = "Yellow (Review):"
    ws[f"B{row}"] = summary.get("yellow_count", 0)
    ws[f"B{row}"].fill = YELLOW_FILL
    row += 1
    ws[f"A{row}"] = "Red (Action):"
    ws[f"B{row}"] = summary.get("red_count", 0)
    ws[f"B{row}"].fill = RED_FILL
    row += 1
    ws[f"A{row}"] = "Total Variance:"
    ws[f"B{row}"] = summary.get("total_variance", 0)
    ws[f"B{row}"].number_format = CURRENCY_FORMAT
    
    # Processor details table
    row += 2
    ws[f"A{row}"] = "Processor Status Details"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    # Headers
    headers = ["Processor", "Status", "SPI Gross", "Proc Gross", "Variance", "Variance %", "Top Reason"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    # Data rows
    row += 1
    for status_dict in daily_statuses:
        ws.cell(row=row, column=1, value=status_dict.get("processor", ""))
        
        status_val = status_dict.get("status", "")
        status_cell = ws.cell(row=row, column=2, value=status_val.upper())
        if status_val == "green":
            status_cell.fill = GREEN_FILL
        elif status_val == "yellow":
            status_cell.fill = YELLOW_FILL
        else:
            status_cell.fill = RED_FILL
        
        ws.cell(row=row, column=3, value=status_dict.get("spi_target_gross", 0)).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=4, value=status_dict.get("proc_target_gross", 0)).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=5, value=status_dict.get("variance_amount", 0)).number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=6, value=status_dict.get("variance_pct", 0) / 100).number_format = PERCENT_FORMAT
        ws.cell(row=row, column=7, value=status_dict.get("top_reason_code", ""))
        
        # Apply borders
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
        
        row += 1
    
    # Auto-width columns
    _auto_width(ws)


# =============================================================================
# Exceptions Sheet
# =============================================================================

def _create_exceptions_sheet(wb: Workbook, exceptions_df: pd.DataFrame, meta: Dict):
    """Create exceptions sheet with variance breakdown by reason code"""
    ws = wb.create_sheet("Exceptions")
    
    # Title
    ws["A1"] = "Exception Breakdown by Reason Code"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Date: {meta.get('target_day', '')}"
    
    row = 4
    
    # Headers
    headers = ["Date", "Processor", "Reason Code", "Amount", "Direction", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    
    # Data rows
    row += 1
    if not exceptions_df.empty:
        for _, exc_row in exceptions_df.iterrows():
            ws.cell(row=row, column=1, value=str(exc_row.get("date", "")))
            ws.cell(row=row, column=2, value=exc_row.get("processor", ""))
            ws.cell(row=row, column=3, value=exc_row.get("reason_code", ""))
            ws.cell(row=row, column=4, value=exc_row.get("amount", 0)).number_format = CURRENCY_FORMAT
            ws.cell(row=row, column=5, value=exc_row.get("direction", ""))
            ws.cell(row=row, column=6, value=exc_row.get("status", ""))
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = THIN_BORDER
            
            row += 1
    else:
        ws.cell(row=row, column=1, value="No exceptions found")
    
    # Reason code legend
    row += 2
    ws[f"A{row}"] = "Reason Code Legend"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    legend = [
        ("WITHIN_TOLERANCE", "Variance within acceptable limits"),
        ("TIMING_CUTOFF", "Event recorded in different days between systems"),
        ("REFUND_FAILURE", "Refund reversal in SPI (positive adjustment)"),
        ("PROCESSOR_ONLY", "Transaction in processor, not in SPI"),
        ("SPI_ONLY", "Transaction in SPI, not in processor"),
        ("DISPUTE_LIFECYCLE", "Chargeback/dispute timing differences"),
        ("FEE_VARIANCE", "Processing fee differences"),
        ("UNEXPLAINED", "Requires investigation"),
    ]
    
    for code, description in legend:
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=description)
        row += 1
    
    _auto_width(ws)


# =============================================================================
# Bridge Sheet (Month-End Style)
# =============================================================================

def _create_bridge_sheet(wb: Workbook, daily_statuses: List[Dict], meta: Dict):
    """Create bridge table sheet in legacy format"""
    ws = wb.create_sheet("Bridge")
    
    entity_name = meta.get("entity", "Unknown")
    target_day = meta.get("target_day", "")
    
    # Title
    ws["A1"] = f"Reconciliation Bridge - {entity_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Period: {target_day}"
    
    row = 4
    
    # Calculate totals from daily statuses
    total_spi_charges = sum(s.get("spi_charge_gross", 0) for s in daily_statuses)
    total_spi_refunds = sum(s.get("spi_refund_gross", 0) for s in daily_statuses)
    total_spi_refund_failures = sum(s.get("spi_refund_failure_gross", 0) for s in daily_statuses)
    total_proc_charges = sum(s.get("proc_charge_gross", 0) for s in daily_statuses)
    total_proc_refunds = sum(s.get("proc_refund_gross", 0) for s in daily_statuses)
    total_proc_fees = sum(s.get("proc_fee_amount", 0) for s in daily_statuses)
    total_variance = sum(s.get("variance_amount", 0) for s in daily_statuses)
    
    # Bridge table structure (mimics legacy workbooks)
    bridge_items = [
        ("SPI/CRM Activity", None, True),
        ("Sales (Charges)", total_spi_charges, False),
        ("Refunds", total_spi_refunds, False),
        ("Refund Failures", total_spi_refund_failures, False),
        ("SPI Gross Total", total_spi_charges + total_spi_refunds + total_spi_refund_failures, True),
        ("", None, False),
        ("Processor Activity", None, True),
        ("Sales (Charges)", total_proc_charges, False),
        ("Refunds", total_proc_refunds, False),
        ("Processing Fees", total_proc_fees, False),
        ("Processor Gross Total", total_proc_charges + total_proc_refunds, True),
        ("", None, False),
        ("Reconciliation", None, True),
        ("SPI Gross Total", total_spi_charges + total_spi_refunds + total_spi_refund_failures, False),
        ("Less: Processor Gross Total", -(total_proc_charges + total_proc_refunds), False),
        ("Gross Variance", total_variance, True),
        ("", None, False),
        ("Variance Analysis", None, True),
        ("Timing Cutoff", sum(s.get("variance_breakdown", {}).get("timing_cutoff", 0) for s in daily_statuses), False),
        ("Refund Failures", sum(s.get("variance_breakdown", {}).get("refund_failure", 0) for s in daily_statuses), False),
        ("Disputes", sum(s.get("variance_breakdown", {}).get("disputes", 0) for s in daily_statuses), False),
        ("Fees", sum(s.get("variance_breakdown", {}).get("fees", 0) for s in daily_statuses), False),
        ("Unexplained", sum(s.get("variance_breakdown", {}).get("unexplained", 0) for s in daily_statuses), True),
    ]
    
    for label, amount, is_header in bridge_items:
        cell_a = ws.cell(row=row, column=1, value=label)
        if is_header:
            cell_a.font = Font(bold=True)
        
        if amount is not None:
            cell_b = ws.cell(row=row, column=2, value=amount)
            cell_b.number_format = CURRENCY_FORMAT
            if is_header:
                cell_b.font = Font(bold=True)
        
        row += 1
    
    _auto_width(ws)


# =============================================================================
# Details Sheet
# =============================================================================

def _create_details_sheet(wb: Workbook, summary_df: pd.DataFrame, meta: Dict):
    """Create details sheet with raw summary data"""
    ws = wb.create_sheet("Details")
    
    ws["A1"] = "Detailed Metrics"
    ws["A1"].font = Font(bold=True, size=14)
    
    row = 3
    
    if not summary_df.empty:
        # Headers
        headers = list(summary_df.columns)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        
        # Data
        row += 1
        for _, data_row in summary_df.iterrows():
            for col, header in enumerate(headers, 1):
                value = data_row[header]
                cell = ws.cell(row=row, column=col, value=value)
                if header == "value" and isinstance(value, (int, float)):
                    cell.number_format = CURRENCY_FORMAT
            row += 1
    
    # Files processed
    row += 2
    ws[f"A{row}"] = "Files Processed"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    for file_path in meta.get("files_processed", []):
        ws.cell(row=row, column=1, value=file_path)
        row += 1
    
    _auto_width(ws)


# =============================================================================
# Helpers
# =============================================================================

def _auto_width(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width